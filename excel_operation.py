"""
Excel を操作する簡単な Python スクリプト

方法1: openpyxl（Excel がインストールされていなくても動作）
方法2: xlwings（Excel アプリケーションを直接操作、Windows/Mac対応）
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 出力ファイル名
OUTPUT_FILE = "output.xlsx"


def create_and_write():
    """新しい Excel ブックを作成して書き込む"""
    wb = Workbook()
    ws = wb.active
    ws.title = "サンプル"

    # セルに値を書き込む
    ws["A1"] = "商品名"
    ws["B1"] = "数量"
    ws["C1"] = "単価"
    ws["D1"] = "合計"

    # 見出しを太字に
    for col in ["A1", "B1", "C1", "D1"]:
        ws[col].font = Font(bold=True)

    # データを書き込む
    data = [
        ["りんご", 5, 120],
        ["みかん", 10, 80],
        ["バナナ", 3, 150],
    ]
    for i, row in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
        ws.cell(row=i, column=3, value=row[2])
        ws.cell(row=i, column=4, value=row[1] * row[2])  # 合計

    # 列幅を調整
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10

    wb.save(OUTPUT_FILE)
    print(f"✅ {OUTPUT_FILE} を作成しました。")


def read_excel(file_path):
    """既存の Excel ファイルを読み込んで内容を表示"""
    wb = load_workbook(file_path, read_only=False)
    ws = wb.active
    print(f"\n📖 シート名: {ws.title}\n")

    for row in ws.iter_rows(min_row=1, values_only=True):
        print(row)

    wb.close()


def update_existing_excel(file_path):
    """既存の Excel ファイルを開いて更新する"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 既存のデータを読み込む
    print("既存のデータ:")
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(row)
    
    # 新しいデータを追加
    last_row = ws.max_row
    ws.cell(row=last_row + 1, column=1, value="ぶどう")
    ws.cell(row=last_row + 1, column=2, value=8)
    ws.cell(row=last_row + 1, column=3, value=200)
    ws.cell(row=last_row + 1, column=4, value=8 * 200)
    
    # 保存
    wb.save(file_path)
    print(f"\n✅ {file_path} を更新しました。")


def read_specific_range(file_path, sheet_name=None, start_cell="A1", end_cell="D10"):
    """特定の範囲のセルを読み込む"""
    wb = load_workbook(file_path, read_only=True)
    
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # 範囲を指定して読み込み
    for row in ws[start_cell:end_cell]:
        row_values = [cell.value for cell in row]
        print(row_values)
    
    wb.close()


# ===== xlwings を使った方法（Excel アプリケーションを直接操作） =====
def excel_with_xlwings():
    """
    xlwings を使った Excel 操作
    Excel アプリケーションがインストールされている必要があります
    Windows/Mac 両方対応
    """
    try:
        import xlwings as xw
        
        # Excel を起動（既に開いている場合はそれを使用）
        app = xw.App(visible=True, add_book=False)
        
        # 新しいブックを作成、または既存のブックを開く
        wb = app.books.add()  # 新規作成
        # wb = app.books.open('existing_file.xlsx')  # 既存ファイルを開く
        
        ws = wb.sheets[0]
        ws.name = "xlwingsサンプル"
        
        # セルに値を書き込む
        ws.range('A1').value = 'Hello'
        ws.range('B1').value = 'World'
        ws.range('A2').value = 100
        ws.range('B2').value = 200
        ws.range('C2').value = '=A2+B2'  # 数式も設定可能
        
        # 範囲に値を一括で書き込む
        data = [
            ['商品名', '数量', '単価'],
            ['りんご', 5, 120],
            ['みかん', 10, 80],
        ]
        ws.range('A4').value = data
        
        # セルの値を読み込む
        value = ws.range('A1').value
        print(f"A1の値: {value}")
        
        # 範囲を読み込む
        range_data = ws.range('A4:C6').value
        print(f"範囲のデータ: {range_data}")
        
        # 保存
        wb.save('xlwings_output.xlsx')
        print("✅ xlwings_output.xlsx を作成しました。")
        
        # 閉じる（Excel を残す場合は wb.close() だけ）
        wb.close()
        app.quit()
        
    except ImportError:
        print("⚠️  xlwings がインストールされていません。")
        print("   インストール: pip install xlwings")
    except Exception as e:
        print(f"⚠️  xlwings の実行中にエラー: {e}")


if __name__ == "__main__":
    print("=" * 50)
    print("方法1: openpyxl を使った操作")
    print("=" * 50)
    
    # 1. 新規作成して保存
    create_and_write()
    
    # 2. 保存したファイルを読み込んで表示
    read_excel(OUTPUT_FILE)
    
    # 3. 既存ファイルを更新
    print("\n" + "=" * 50)
    print("既存ファイルの更新")
    print("=" * 50)
    update_existing_excel(OUTPUT_FILE)
    
    # 4. 特定範囲を読み込む
    print("\n" + "=" * 50)
    print("特定範囲の読み込み")
    print("=" * 50)
    read_specific_range(OUTPUT_FILE, start_cell="A1", end_cell="D5")
    
    # 5. xlwings を使った方法（オプション）
    print("\n" + "=" * 50)
    print("方法2: xlwings を使った操作（Excel アプリケーションを直接操作）")
    print("=" * 50)
    excel_with_xlwings()
